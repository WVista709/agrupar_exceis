import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import os
import threading
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

# Constantes
ABAS_EXCEL = [
    "COMPRAS SEFAZ", "COMPRAS ALTERDATA", "COMPRAS PRODUTOS",
    "VENDAS SEFAZ", "VENDAS ALTERDATA", "VENDAS PRODUTOS",
    "SERVIÇOS PRESTADOS", "SERVIÇOS TOMADOS", "CHECK", "PRODUTOS", "APURAÇÃO"
]

ABAS_ARQUIVOS = [
    "COMPRAS SEFAZ", "COMPRAS ALTERDATA", "COMPRAS PRODUTOS",
    "VENDAS SEFAZ", "VENDAS ALTERDATA", "VENDAS PRODUTOS",
    "SERVIÇOS PRESTADOS", "SERVIÇOS TOMADOS"
]

# Variáveis globais de interface
arquivos_selecionados = {aba: None for aba in ABAS_ARQUIVOS}
carregando_popup = None
barra_progresso = None
caixa_nome_excel = None
botao_gerar_excel = None

def mostrar_carregando():
    global carregando_popup, barra_progresso
    carregando_popup = tk.Toplevel()
    carregando_popup.title("Carregando")
    carregando_popup.geometry("300x100")
    carregando_popup.resizable(False, False)
    carregando_popup.transient()
    carregando_popup.grab_set()
    tk.Label(carregando_popup, text="Gerando Excel, aguarde...").pack(pady=10)
    barra_progresso = ttk.Progressbar(carregando_popup, mode="indeterminate", length=250)
    barra_progresso.pack(pady=5)
    barra_progresso.start(10)

def fechar_carregando():
    global carregando_popup, barra_progresso
    if carregando_popup:
        barra_progresso.stop()
        carregando_popup.destroy()
        carregando_popup = None

def selecionar_arquivo(aba_nome, var_label):
    caminho = filedialog.askopenfilename(
        title=f'Escolha o arquivo {aba_nome}.xlsx',
        filetypes=[("Arquivos do Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
    )
    if caminho:
        arquivos_selecionados[aba_nome] = caminho
        var_label.set(os.path.basename(caminho))

def criar_botoes_arquivo(frame, abas, linha_inicial=0):
    for i, aba in enumerate(abas):
        var_label = tk.StringVar(value="Nada selecionado")
        btn = tk.Button(frame, text=aba, width=20, command=lambda a=aba, v=var_label: selecionar_arquivo(a, v))
        btn.grid(row=linha_inicial + i, column=0, padx=5, pady=2)
        lbl = tk.Label(frame, textvariable=var_label, anchor="w", width=40)
        lbl.grid(row=linha_inicial + i, column=1, padx=5, pady=2)

def criar_abas_excel(wb):
    for aba in ABAS_EXCEL:
        wb.create_sheet(aba)

def copiar_planilhas(aba_nome, wb):
    arquivo = arquivos_selecionados.get(aba_nome)
    if arquivo:
        try:
            excel_origem = load_workbook(arquivo)
            aba_excel_origem = excel_origem.active
            aba_destino = wb[aba_nome]
            for row in aba_excel_origem.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        aba_destino[cell.coordinate].value = cell.value
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao copiar planilha {aba_nome}: {e}")

def formatar_valores_contabil(aba, colunas, linhas):
    formato_moeda = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'
    for col in colunas:
        for linha in linhas:
            celula = aba[f"{col}{linha}"]
            celula.number_format = formato_moeda

def adicionar_formula_procv(wb, aba_destino, aba_matriz, valor_procurado, coluna_busca, rotulo):
    aba = wb[aba_destino]
    coluna_destino = aba.max_column + 1
    aba.cell(row=1, column=coluna_destino).value = rotulo
    for linha in range(2, aba.max_row + 1):
        celula_ref = f"{valor_procurado}{linha}"
        formula = f'=IFERROR(VLOOKUP({celula_ref}, \'{aba_matriz}\'!{coluna_busca}:{coluna_busca}, 1, FALSE), "ERRO")'
        aba.cell(row=linha, column=coluna_destino).value = formula

def adicionar_formula_cancelada(wb, aba_destino, coluna_ref, rotulo):
    aba = wb[aba_destino]
    coluna_destino = aba.max_column + 1
    aba.cell(row=1, column=coluna_destino).value = rotulo
    for linha in range(2, aba.max_row + 1):
        celula_ref = f"{coluna_ref}{linha}"
        formula = f'=IF({celula_ref}="AUTORIZADA", "NÃO", "SIM")'
        aba.cell(row=linha, column=coluna_destino).value = formula

def formula_check(wb, celula, formula):
    aba = wb["CHECK"]
    aba[celula].value = f"={formula}"

def formula_generica(wb, aba_nome, celula, formula):
    aba = wb[aba_nome]
    aba[celula].value = f"={formula}"

def formula_somases(wb, aba, linha, coluna, aba_coluna, aba_matriz, aba_matriz_coluna, aba_check):
    aba_check_sheet = wb["CHECK"]
    formula = f'=SUMIFS(\'{aba_matriz}\'!{aba_coluna}:{aba_coluna}, \'{aba_matriz}\'!{aba_matriz_coluna}:{aba_matriz_coluna}, \'{aba}\'!{aba_check})'
    aba_check_sheet.cell(row=linha, column=coluna).value = formula

def tabela_check(wb, mesclar_inicio, mesclar_final, rotulo, cabecalho_linha, linha):
    aba = wb["CHECK"]
    rotulos_linhas = ["NÃO", "SIM", "TOTAL"]
    rotulos_cabecalho = ["CANCELADAS", "SEFAZ", "ALTERDATA", "PRODUTO", "DESCONTO", "ST Ñ APROVEITADO", "ST APROVEITADO"]

    aba.merge_cells(f"{mesclar_inicio}:{mesclar_final}")
    aba[mesclar_inicio].value = rotulo
    aba[mesclar_inicio].alignment = Alignment(horizontal="center", vertical="center")

    for col, texto in enumerate(rotulos_cabecalho, start=1):
        aba.cell(row=cabecalho_linha, column=col).value = texto

    for i, texto in enumerate(rotulos_linhas):
        aba.cell(row=linha + i, column=1).value = texto

    # Fórmulas e somas
    formula_check(wb, "B6", "B5-C5")
    formula_check(wb, "D6", "C5-D5")
    formula_check(wb, "F6", "D6+E5-F5-G5")
    formula_check(wb, "B13", "B12-C12")
    formula_check(wb, "D13", "C12-D12")
    formula_check(wb, "F13", "D13+E12-F12-G12")

    abas = ["COMPRAS", "VENDAS"]
    linhas_iniciais = [3, 10]
    abas_sefaz = ["COMPRAS SEFAZ", "VENDAS SEFAZ"]
    abas_alterdata = ["COMPRAS ALTERDATA", "VENDAS ALTERDATA"]
    abas_produtos = ["COMPRAS PRODUTOS", "VENDAS PRODUTOS"]
    colunas_iniciais_check = [2, 2]

    for i in range(len(abas)):
        linha_check = linhas_iniciais[i]
        coluna_check = colunas_iniciais_check[i]
        celula_referencia_1 = f"A{linha_check}"
        celula_referencia_2 = f"A{linha_check + 1}"

        # SEFAZ
        formula_somases(wb, "CHECK", linha_check, coluna_check, 'P', abas_sefaz[i], 'Y', celula_referencia_1)
        formula_somases(wb, "CHECK", linha_check + 1, coluna_check, 'P', abas_sefaz[i], 'Y', celula_referencia_2)
        # ALTERDATA
        formula_somases(wb, "CHECK", linha_check, coluna_check + 1, 'J', abas_alterdata[i], 'I', celula_referencia_1)
        formula_somases(wb, "CHECK", linha_check + 1, coluna_check + 1, 'J', abas_alterdata[i], 'I', celula_referencia_2)
        # PRODUTOS
        for j, col in enumerate(['M', 'N', 'O', 'P'], start=2):
            formula_somases(wb, "CHECK", linha_check, coluna_check + j, col, abas_produtos[i], 'I', celula_referencia_1)
            formula_somases(wb, "CHECK", linha_check + 1, coluna_check + j, col, abas_produtos[i], 'I', celula_referencia_2)

    colunas_soma = ['B', 'C', 'D', 'E', 'F', 'G']
    for coluna in colunas_soma:
        formula_check(wb, f"{coluna}5", f"SUM({coluna}3:{coluna}4)")
        formula_check(wb, f"{coluna}12", f"SUM({coluna}10:{coluna}11)")

    # ALUGUEL e ENERGIA
    aba.merge_cells("A15:A16")
    aba.cell(row=15, column=1).value = "ALUGUEL"
    aba.cell(row=16, column=2).value = 0
    aba.cell(row=15, column=2).value = "VALOR"
    aba.cell(row=15, column=3).value = "PIS"
    aba.cell(row=15, column=4).value = "COFINS"
    aba.cell(row=16, column=3).value = "=b16 * 1.65%"
    aba.cell(row=16, column=4).value = "=b16 * 7.6%"

    aba.cell(row=17, column=1).value = "ENERGIA"
    aba.cell(row=17, column=2).value = "=SUMIFS('COMPRAS ALTERDATA'!J:J,'COMPRAS ALTERDATA'!H:H,1253)"
    aba.cell(row=17, column=3).value = "=b17 * 1.65%"
    aba.cell(row=17, column=4).value = "=b17 * 7.6%"

    # ERROS
    rotulos_erros = ["ERROS", "SEFAZ", "ALTERDATA", "PRODUTOS"]
    rotulos_erros_linhas = ["COMPRAS", "VENDAS"]
    for col, texto in enumerate(rotulos_erros, start=1):
        aba.cell(row=19, column=col).value = texto
    for i, texto in enumerate(rotulos_erros_linhas):
        aba.cell(row=20 + i, column=1).value = texto

    coluna = ['W', 'Q', 'AA']
    coluna_celula = ['B', 'C', 'D']
    abas_compras = ["COMPRAS SEFAZ", "COMPRAS ALTERDATA", "COMPRAS PRODUTOS"]
    abas_vendas = ["VENDAS SEFAZ", "VENDAS ALTERDATA", "VENDAS PRODUTOS"]
    for aba_compra, aba_venda, col, celula in zip(abas_compras, abas_vendas, coluna, coluna_celula):
        formula_check(wb, f"{celula}20", f'COUNTIF(\'{aba_compra}\'!{col}:{col}, "ERRO")')
        formula_check(wb, f"{celula}21", f'COUNTIF(\'{aba_venda}\'!{col}:{col}, "ERRO")')

    linhas = [3, 4, 5, 6, 10, 11, 12, 13, 16, 17]
    formatar_valores_contabil(aba, colunas_soma, linhas)

def tabela_produtos(wb):
    rotulos = ["NOME", "ORIGEM", "COMPRAS", "PIS", "COFINS", "ICMS", "", "VENDAS BRUTAS", "VENDAS 6929", "VENDAS 5929", "VENDAS LIQUIDAS", "PIS", "COFINS", "ICMS"]
    aba_origem_compras = wb["COMPRAS PRODUTOS"]
    aba_origem_vendas = wb["VENDAS PRODUTOS"]
    aba_destino = wb["PRODUTOS"]

    produtos_compras = {cell.value for cell in aba_origem_compras['F'][1:] if cell.value}
    produtos_vendas = {cell.value for cell in aba_origem_vendas['F'][1:] if cell.value}
    todos_produtos = produtos_compras.union(produtos_vendas)

    for i, produto in enumerate(sorted(todos_produtos), start=2):
        aba_destino.cell(row=i, column=1).value = produto
        origem = "AMBOS" if produto in produtos_compras and produto in produtos_vendas else \
            "COMPRAS" if produto in produtos_compras else "VENDAS"
        aba_destino.cell(row=i, column=2).value = origem

        # COMPRAS
        formula_generica(wb, "PRODUTOS", f"C{i}", f"SUMIFS('COMPRAS PRODUTOS'!M:M,'COMPRAS PRODUTOS'!F:F,A{i})")
        formula_generica(wb, "PRODUTOS", f"D{i}", f"SUMIFS('COMPRAS PRODUTOS'!T:T,'COMPRAS PRODUTOS'!F:F,A{i})")
        formula_generica(wb, "PRODUTOS", f"E{i}", f"SUMIFS('COMPRAS PRODUTOS'!U:U,'COMPRAS PRODUTOS'!F:F,A{i})")
        formula_generica(wb, "PRODUTOS", f"F{i}", f"SUMIFS('COMPRAS PRODUTOS'!V:V,'COMPRAS PRODUTOS'!F:F,A{i})")

        # VENDAS
        formula_generica(wb, "PRODUTOS", f"H{i}", f"SUMIFS('VENDAS PRODUTOS'!M:M,'VENDAS PRODUTOS'!F:F,A{i})")
        formula_generica(wb, "PRODUTOS", f"I{i}", f"SUMIFS('VENDAS PRODUTOS'!M:M,'VENDAS PRODUTOS'!F:F,A{i}, 'VENDAS PRODUTOS'!G:G, 5929)")
        formula_generica(wb, "PRODUTOS", f"J{i}", f"SUMIFS('VENDAS PRODUTOS'!M:M,'VENDAS PRODUTOS'!F:F,A{i}, 'VENDAS PRODUTOS'!G:G, 6929)")
        formula_generica(wb, "PRODUTOS", f"K{i}", f"H{i}-I{i}-J{i}")
        formula_generica(wb, "PRODUTOS", f"L{i}", f"SUMIFS('VENDAS PRODUTOS'!T:T,'VENDAS PRODUTOS'!F:F,A{i})")
        formula_generica(wb, "PRODUTOS", f"M{i}", f"SUMIFS('VENDAS PRODUTOS'!U:U,'VENDAS PRODUTOS'!F:F,A{i})")
        formula_generica(wb, "PRODUTOS", f"N{i}", f"SUMIFS('VENDAS PRODUTOS'!V:V,'VENDAS PRODUTOS'!F:F,A{i})")

    for j, rotulo in enumerate(rotulos, start=1):
        aba_destino.cell(row=1, column=j).value = rotulo

    colunas = ['C', 'D', 'E', 'F', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
    linhas = range(2, aba_destino.max_row + 1)
    formatar_valores_contabil(aba_destino, colunas, linhas)

def tabela_apuracao(wb):
    aba = wb["APURAÇÃO"]
    rotulos_horizontais = ["TIPO", "VALOR", "PIS", "COFINS", "ICMS"]
    rotulos_verticais = ["COMPRAS", "VENDAS", "TOTAL", "TIPO"]

    for col, titulo in enumerate(rotulos_horizontais, start=1):
        aba.cell(row=1, column=col).value = titulo
    for lin, tipo in enumerate(rotulos_verticais, start=2):
        aba.cell(row=lin, column=1).value = tipo

    linha_compras_apuracao = ['C', 'D', 'E', 'F']
    linha_vendas_apuracao = ['K', 'L', 'M', 'N']
    linha_apuracao = ['B', 'C', 'D', 'E']

    formatar_valores_contabil(aba, linha_apuracao, [2, 3, 4])

    for col_ap, col_compra, col_venda in zip(linha_apuracao, linha_compras_apuracao, linha_vendas_apuracao):
        if col_ap in ['C', 'D']:
            formula = f'SUM(PRODUTOS!{col_compra}:{col_compra}) + SUM(CHECK!{col_ap}16, CHECK!{col_ap}17)'
        else:
            formula = f'SUM(PRODUTOS!{col_compra}:{col_compra})'
        formula_generica(wb, "APURAÇÃO", f'{col_ap}2', formula)
        formula_generica(wb, "APURAÇÃO", f'{col_ap}3', f'-SUM(PRODUTOS!{col_venda}:{col_venda})')
        formula_generica(wb, "APURAÇÃO", f'{col_ap}4', f'{col_ap}2+{col_ap}3')
        formula_generica(wb, "APURAÇÃO", f'{col_ap}5', f'IF({col_ap}4>=0, "CRÉDITO", "DÉBITO")')
    formula_generica(wb, "APURAÇÃO", 'B5', f'IF(B4>=0, "FATUROU MENOS", "FATUROU MAIS")')

def criar_excel():
    global carregando_popup, barra_progresso, botao_gerar_excel
    nome_arquivo = caixa_nome_excel.get().strip()
    diretorio = filedialog.askdirectory()
    if not (diretorio and nome_arquivo):
        messagebox.showwarning("AVISO", "Por favor, escolha um nome e um diretório válidos para o arquivo.")
        botao_gerar_excel.config(state="normal")
        return
    caminho_arquivo = os.path.join(diretorio, f'{nome_arquivo}.xlsx')
    if os.path.exists(caminho_arquivo):
        resposta = messagebox.askyesno("Arquivo Existente", f"O arquivo '{nome_arquivo}.xlsx' já existe. Deseja substituir?")
        if not resposta:
            botao_gerar_excel.config(state="normal")
            return
    try:
        mostrar_carregando()
        wb = Workbook()
        wb.remove(wb.active)
        criar_abas_excel(wb)
        for aba in arquivos_selecionados:
            copiar_planilhas(aba, wb)
        abas_formulas = ["COMPRAS", "VENDAS"]
        for aba in abas_formulas:
            adicionar_formula_procv(wb, f'{aba} SEFAZ', f"{aba} ALTERDATA", 'C', 'C', "ALTERDATA")
            adicionar_formula_procv(wb, f'{aba} SEFAZ', f"{aba} PRODUTOS", 'C', 'B', "PRODUTOS")
            adicionar_formula_cancelada(wb, f'{aba} SEFAZ', "N", "CANCELADAS")
            adicionar_formula_procv(wb, f'{aba} ALTERDATA', f"{aba} SEFAZ", 'C', 'C', "SEFAZ")
            adicionar_formula_procv(wb, f'{aba} ALTERDATA', f"{aba} PRODUTOS", 'C', 'B', "PRODUTOS")
            adicionar_formula_procv(wb, f'{aba} PRODUTOS', f"{aba} SEFAZ", 'B', 'C', "SEFAZ")
            adicionar_formula_procv(wb, f'{aba} PRODUTOS', f"{aba} ALTERDATA", 'B', 'C', "ALTERDATA")
        tabela_check(wb, "A1", "G1", "COMPRAS", 2, 3)
        tabela_check(wb, "A8", "G8", "VENDAS", 9, 10)
        tabela_produtos(wb)
        tabela_apuracao(wb)
        wb.save(caminho_arquivo)
        messagebox.showinfo("Sucesso", f"Arquivo '{nome_arquivo}.xlsx' criado com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao criar o arquivo: {e}")
    finally:
        fechar_carregando()
        botao_gerar_excel.config(state="normal")

def iniciar_thread_criar_excel():
    global botao_gerar_excel
    botao_gerar_excel.config(state="disabled")
    threading.Thread(target=criar_excel, daemon=True).start()

def botao_para_agrupar_exceis(janela, x, y):
    global caixa_nome_excel, botao_gerar_excel
    tk.Label(janela, text="Nome do novo Excel:").place(x=x, y=y)
    caixa_nome_excel = tk.Entry(janela, width=30)
    caixa_nome_excel.place(x=x + 130, y=y)
    botao_gerar_excel = tk.Button(janela, text="Gerar Excel", command=iniciar_thread_criar_excel)
    botao_gerar_excel.place(x=x + 130, y=y + 30)

def interface():
    janela = tk.Tk()
    janela.title("Interface de Agrupamento de Excel")
    janela.geometry("600x600")

    frame_compras = tk.LabelFrame(janela, text="COMPRAS", padx=10, pady=10)
    frame_compras.place(x=20, y=20)
    criar_botoes_arquivo(frame_compras, ["COMPRAS SEFAZ", "COMPRAS ALTERDATA", "COMPRAS PRODUTOS"])

    frame_vendas = tk.LabelFrame(janela, text="VENDAS", padx=10, pady=10)
    frame_vendas.place(x=20, y=160)
    criar_botoes_arquivo(frame_vendas, ["VENDAS SEFAZ", "VENDAS ALTERDATA", "VENDAS PRODUTOS"])

    frame_servico = tk.LabelFrame(janela, text="SERVIÇOS", padx=10, pady=10)
    frame_servico.place(x=20, y=300)
    criar_botoes_arquivo(frame_servico, ["SERVIÇOS PRESTADOS", "SERVIÇOS TOMADOS"])

    botao_para_agrupar_exceis(janela, 150, 500)
    janela.mainloop()

if __name__ == "__main__":
    interface()